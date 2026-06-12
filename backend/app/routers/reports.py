"""Dashboard and reporting endpoints."""

import csv
import io
from datetime import datetime
from typing import List

from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import User, Internship
from app.schemas import DashboardStats, AgreementStatusItem, FinalReportItem
from app.auth import get_current_active_user, require_any_staff, require_role
from app.services import reports as _reports
from app.services.report_pdf import generate_final_report_pdf

router = APIRouter(prefix="/internships", tags=["reports"])


def _row_data(i: Internship) -> list:
    """Build a flat list of exportable fields for an internship."""
    return [
        f"{i.student.first_name} {i.student.last_name}" if i.student else "",
        i.student.email if i.student else "",
        i.company.name if i.company else "",
        i.company.sector if i.company else "",
        i.status,
        str(i.start_date) if i.start_date else "",
        str(i.end_date) if i.end_date else "",
        f"{i.teacher.first_name} {i.teacher.last_name}" if i.teacher else "",
        f"{i.mentor.first_name} {i.mentor.last_name}" if i.mentor else "",
        i.proposal.status if i.proposal else "",
        i.agreement.status if i.agreement else "",
    ]


@router.get("/stats/dashboard", response_model=DashboardStats)
def get_dashboard_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    return _reports.get_dashboard_stats(db, current_user)


@router.get("/reports/agreements", response_model=List[AgreementStatusItem])
def get_agreement_status_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_any_staff),
):
    return _reports.get_agreement_status_report(db, current_user)


@router.get("/{internship_id}/final-report", response_model=FinalReportItem)
def get_final_report(
    internship_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    return _reports.get_final_report(db, current_user, internship_id)


@router.get("/{internship_id}/final-report/pdf")
def get_final_report_pdf(
    internship_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    pdf_path, student = generate_final_report_pdf(db, current_user, internship_id)
    filename = f"eindrapport_{student.first_name}_{student.last_name}.pdf"
    return FileResponse(
        str(pdf_path),
        media_type="application/pdf",
        filename=filename,
    )


@router.get("/reports/export/csv")
def export_internships_csv(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["teacher", "committee", "admin"])),
):

    internships = _reports._apply_role_filter(db.query(Internship), current_user).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Student",
            "Email",
            "Bedrijf",
            "Sector",
            "Status",
            "Startdatum",
            "Einddatum",
            "Docent",
            "Mentor",
            "Voorstel Status",
            "Overeenkomst Status",
        ]
    )
    for i in internships:
        writer.writerow(_row_data(i))

    output.seek(0)
    filename = f"stage_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/reports/export/excel")
def export_internships_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(["teacher", "committee", "admin"])),
):

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    internships = _reports._apply_role_filter(db.query(Internship), current_user).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stage Overzicht"

    # Header
    headers = [
        "Student",
        "Email",
        "Bedrijf",
        "Sector",
        "Status",
        "Startdatum",
        "Einddatum",
        "Docent",
        "Mentor",
        "Voorstel Status",
        "Overeenkomst Status",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="00798C", end_color="00798C", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data rows
    for row, i in enumerate(internships, 2):
        for col, val in enumerate(_row_data(i), 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None), default=0
        )
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"stage_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
